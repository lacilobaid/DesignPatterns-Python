import openpyxl
import time
import pypyodbc
from openpyxl import Workbook,load_workbook

class excel_db:
    def __init__(self):
        self.workbook = load_workbook("urun_bilgi.xlsx")
        self.sheet = self.workbook.active
    def veri_oku(self):
        for satir in range(1, self.sheet.max_row + 1):
            for sutun in range(1, self.sheet.max_column + 1):
                print(" | " + str(self.sheet.cell(satir, sutun).value) + " | ", end="")
            print()
    def veri_ekle(self):
        urun_id = int(input('Urun id giriniz: '))
        urun_adi = input('Urun adi giriniz: ')
        urun_fiyat = int(input('Urun fiyati giriniz: '))
        urun_stok = int(input('Urun stok giriniz: '))
        self.sheet.append([urun_id, urun_adi, urun_fiyat, urun_stok])


class mssql_db:
    def  __init__(self):
        yeni_database = pypyodbc.connect(
            'Driver={SQL Server};'
            'Server=DESKTOP-L5EBS4A\SQLEXPRESS;'
            'Database=urun_bilgi;'
            'Trusted_Connection=True;'
        )
        self.imlec = yeni_database.cursor()
    def urun_goster(self):
        self.imlec.execute('Select * from urun_bilgi_t')
        urunler = self.imlec.fetchall()
        if len(urunler) == 0:
            print("Urun bulunmamaktadır.")
        else:
            for i in urunler:
                print(i)
    def urun_ekle(self):
        urun_id = int(input('Urun id giriniz: '))
        urun_adi = input('Urun adi giriniz: ')
        urun_fiyat = int(input('Urun fiyati giriniz: '))
        urun_stok = int(input('Urun stok giriniz: '))
        sorgu = 'Insert into urun_bilgi_t (urunID, urunAdi, urunFiyat, urunStok) Values (?,?,?,?)'
        self.imlec.execute(sorgu,(urun_id,urun_adi,urun_fiyat,urun_stok))
        self.imlec.commit()

class db_adapter(excel_db):
    new_db = mssql_db()
    def veri_oku(self):
        self.new_db.urun_goster()
    def veri_ekle(self):
        self.new_db.urun_ekle()

if __name__ == '__main__':
    excel_database=excel_db()
    adapter=db_adapter()
    print('* URUN * BILGI * SISTEMI *')
    while True:
        print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
        secim = input("Cikis icin Q, \nExcel islemleri icin E, \nMSSQL islemleri icin S'ye basiniz: ")
        if secim == 'Q' or secim == 'q':
            print('Program sonlandiriliyor...')
            time.sleep(2)
            break
        elif secim == 'E' or secim == 'e':
            islem = input('1. Urunleri goster \n2. Urun ekle \nBir secim yapiniz: ')
            if islem == '1':
                excel_database.veri_oku()
            elif islem=='2':
                excel_database.veri_ekle()
                print('Urun ekleniyor...')
                time.sleep(2)
                print('Urun eklendi...')
            else: print('Hatali islem secimi...')
        elif secim == 'S' or secim == 's':
            islem = input('1. Urunleri goster \n2. Urun ekle \nBir secim yapiniz: ')
            if islem == '1':
                adapter.veri_oku()
            elif islem=='2':
                adapter.veri_ekle()
                print('Urun ekleniyor...')
                time.sleep(2)
                print('Urun eklendi...')
            else: print('Hatali islem secimi...')
        else: print('Hatali secim yaptiniz...')


